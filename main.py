import base64
import asyncio
import aiohttp
import aiosqlite
from datetime import datetime
import pytz
import sys
import pandas as pd
import random
import time
import sqlite3
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook

if sys.platform.startswith("win"):
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

DB_NAME = "Instagram.db"
EXCEL_FILE = "InstagramData.xlsx"


def encode_auth(username: str) -> str:
    raw = f"-1::{username}::rJP2tBRKf6ktbRqPUBtRE9klgBWb7d"
    base64_str = base64.b64encode(raw.encode()).decode()
    return base64_str.replace("+", "-").replace("/", "_").rstrip("=")


async def fetch_story(session: aiohttp.ClientSession, username: str) -> dict | None:
    encoded_auth = encode_auth(username)
    url = "https://anonstories.com/api/v1/story"
    headers = {
        "accept": "*/*",
        "content-type": "application/x-www-form-urlencoded; charset=UTF-8",
        "origin": "https://anonstories.com",
        "referer": "https://anonstories.com/",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
        "x-requested-with": "XMLHttpRequest",
    }
    data = {"auth": encoded_auth}
    async with session.post(url, headers=headers, data=data) as resp:
        if resp.status == 200:
            return await resp.json()
        return None


async def init_db():
    async with aiosqlite.connect(DB_NAME) as db:
        await db.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT,
                full_name TEXT,
                is_private TEXT,
                is_verified TEXT,
                profile_pic_url TEXT,
                follower_count INTEGER,
                following_count INTEGER,
                mention TEXT,
                time TEXT
            )
        """
        )
        await db.execute(
            """
            CREATE TABLE IF NOT EXISTS users_invalid (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT,
                error_message TEXT,
                time TEXT
            )
        """
        )
        await db.commit()


async def save_user_data(username: str, data: dict | None):
    tz_iran = pytz.timezone("Asia/Tehran")
    now = datetime.now(tz_iran).strftime("%Y-%m-%d %H:%M:%S")

    async with aiosqlite.connect(DB_NAME) as db:
        try:
            if not data or not data.get("user_info"):
                raise ValueError("No user_info found")

            user_info = data["user_info"]
            full_name = user_info.get("full_name", "")
            is_private = str(user_info.get("is_private", False))
            is_verified = str(user_info.get("is_verified", False))
            profile_pic_url = user_info.get("profile_pic_url", "")
            follower_count = user_info.get("followers", 0)
            following_count = user_info.get("following", 0)
            stories = data.get("stories", [])

            mentions = list({m for story in stories for m in story.get("mentions", [])})
            if not mentions:
                mentions = [None]

            for mention in mentions:
                await db.execute(
                    """
                    INSERT INTO users (username, full_name, is_private, is_verified,
                    profile_pic_url, follower_count, following_count, mention, time)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                    (
                        username,
                        full_name,
                        is_private,
                        is_verified,
                        profile_pic_url,
                        follower_count,
                        following_count,
                        mention,
                        now,
                    ),
                )

            await db.commit()
            return True

        except Exception as e:
            await db.execute(
                """
                INSERT INTO users_invalid (username, error_message, time)
                VALUES (?, ?, ?)
            """,
                (username, str(e), now),
            )
            await db.commit()
            return False


async def export_excel():
    conn = sqlite3.connect(DB_NAME)

    def save_sheet(df, sheet_name, writer):
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]

        header_font = Font(bold=True)
        alignment = Alignment(horizontal="center", vertical="center")

        for i, column in enumerate(df.columns, 1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = 20
            cell = ws.cell(row=1, column=i)
            cell.font = header_font
            cell.alignment = alignment

    df_users = pd.read_sql("SELECT * FROM users", conn)
    df_invalid = pd.read_sql("SELECT * FROM users_invalid", conn)

    translations = {
        "username": "نام کاربری",
        "full_name": "نام کامل",
        "is_private": "خصوصی",
        "is_verified": "تأیید شده",
        "profile_pic_url": "عکس پروفایل",
        "mention": "منشن",
        "follower_count": "تعداد فالوئر",
        "following_count": "تعداد فالووینگ",
        "time": "زمان دریافت",
        "error_message": "پیام خطا",
    }

    df_users_fa = df_users.rename(columns=translations)
    df_invalid_fa = df_invalid.rename(columns=translations)

    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
        save_sheet(df_users, "Users (EN)", writer)
        save_sheet(df_users_fa, "کاربران (FA)", writer)
        save_sheet(df_invalid, "Invalid (EN)", writer)
        save_sheet(df_invalid_fa, "نامعتبرها (FA)", writer)

    conn.close()


async def process_usernames(usernames: list[str]):
    success, fail = 0, 0
    async with aiohttp.ClientSession() as session:
        for i, username in enumerate(usernames, 1):
            try:
                data = await fetch_story(session, username)
                if await save_user_data(username, data):
                    print(f"✅ {username} saved.")
                    success += 1
                else:
                    print(f"❌ Failed saving {username}")
                    fail += 1
            except Exception as e:
                print(f"❌ Exception for {username}: {e}")
                await save_user_data(username, None)
                fail += 1
            await asyncio.sleep(random.uniform(1, 2))
            print(f"⏳ Remaining: {len(usernames)-i}")

    return success, fail


async def main():
    start = time.time()
    await init_db()
    df = pd.read_excel("1.xlsx", header=None)
    usernames = df.iloc[:, 0].dropna().astype(str).tolist()
    # usernames = [
    #     "shirinderiis",
    #     "saraliife1",
    #     "terazhdi",
    #     "herosheemaz",
    #     "zane_emroozii",
    #     "qome_ziba",
    #     "roya_gilabadi",
    #     "pumkinsara",
    #     "home_healthy_pro",
    #     "teory.x",
    #     "celin_family",
    #     "marjan.kiiaa",
    #     "dollar_ir",
    #     "isensi",
    #     "baranfamily.brt",
    #     "hiddencamprank",
    #     "tvfactt",
    #     "zheen_magazine",
    #     "honarmandanclip1",
    #     "yasiin",
    #     "dr.arezoufarajpour",
    #     "nimaaghadiri",
    #     "2",
    #     "gilan_state",
    #     "mashhad.koja.begardim",
    #     "zibashooo.parastooo",
    #     "samir_samirashirzad",
    #     "mom_hosein_hosna",
    #     "sajjadesbati",
    #     "golawbi",
    #     "bebin_funny",
    #     "chonan_to",
    #     "tiksmag",
    #     "emperaturi.sorkh",
    #     "sanam_samipoor",
    #     "factance",
    #     "soma.yehlife",
    #     "itsbanidalili",
    #     "erfanalirezai",
    #     "start.body",
    #     "samanehpolo",
    #     "khoshkhorak",
    #     "arianabrouni",
    #     "kosar_familly",
    #     "honarmandannet",
    #     "zhatiis",
    #     "irani_dubsmash",
    #     "pardis_abdolmohmdi",
    #     "iranianarchitecturestudents",
    #     "aydamohsen",
    #     "bitreax",
    #     "iranian_acc",
    #     "fact.magt",
    #     "maskaye_khanegi",
    #     "best.clipps",
    #     "psp.persian",
    #     "shahin.shahr",
    #     "ashkan.wichi",
    #     "farhadpaz",
    #     "mohamadaminkarimpor",
    #     "mehdisherafat_",
    #     "salimohh",
    #     "navidghiassi",
    #     "honarstan_idea",
    #     "zeynab_6333",
    #     "hamedgilak",
    #     "fatemeh.chegini",
    #     "becha.nejebad",
    #     "amiirvatankhah",
    #     "paryapartovifard",
    #     "parmis_org4",
    #     "samane_qomi",
    #     "zheste_akkasi",
    #     "thetaranneh",
    #     "saharbeygii_",
    #     "vahiddn",
    #     "ghazall.sadrii",
    #     "melikanikkhah_",
    #     "dr_soroush73",
    #     "big.tasterr_1",
    #     "_raziyehaziz",
    #     "digiato",
    #     "about.farnia",
    #     "fahimebyt",
    #     "lazizz_fpgn",
    #     "mitra_eini",
    #     "rap.bhdnews",
    #     "rasta.adventur",
    #     "_kamaand",
    #     "username",
    #     "ravanshnasie_khanevadeh",
    #     "nazila_parvinghods",
    #     "niloofar",
    #     "rezvan.banooo",
    #     "baranfamily.arak",
    #     "amir_soleymani_food",
    #     "akofood",
    #     "sevin.familly",
    #     "gymclub.ir",
    #     "miladfazeli72",
    #     "_aghoshe_to",
    #     "ashpazio.khanedari",
    #     "shallyzomorodi",
    #     "mazbar_abadan",
    #     "iz.mohsen",
    #     "shaqayqmoqadam",
    #     "hodavvv_",
    #     "hamed.shahanii",
    #     "shayaaanrahaaa",
    #     "car.ir_mag",
    #     "parniya.astaneh",
    #     "dr.siyami",
    #     "agrinkazemi96",
    #     "zahra.newstyle",
    #     "maryam.toosi",
    #     "befunim",
    #     "mouzik_mazandaran",
    #     "kazem.aghlmand",
    #     "faeze.raaad",
    #     "taranehh_melody",
    #     "partotaeb",
    #     "hamsarone",
    #     "vayfarnoosh",
    #     "shahrouz.chorakchi",
    #     "parisa.yazdaniiiii",
    #     "dorfa",
    #     "jorvajor.fact",
    #     "azin_hakhamanesh2019",
    #     "whats.fact",
    #     "mahsa_motahariann",
    #     "ravanshenase_koodak_",
    #     "3pide.diary",
    #     "zibatarin.goll",
    #     "isfahan_university",
    #     "dr_toobashabani",
    #     "alirezaajafarzadeh",
    #     "mitra_greenlife",
    #     "asal.ranjbaar",
    #     "barsam_adrian_dian_shahi",
    #     "sadaf_beauty",
    #     "haniyehhashemi71",
    #     "ahwaz_vlog",
    #     "about_ftm77",
    #     "elmirahashemiii",
    #     "taranenevesht",
    #     "tanznamairan",
    #     "bazidooni",
    #     "zhesteakkasii",
    #     "mrchef.iran",
    #     "mahtab._.lifestory",
    #     "shivashho",
    #     "elahe.mansoriyan",
    #     "urmiye_dangbej",
    #     "sahamsharpi",
    #     "tvsorkh",
    #     "reygram",
    #     "viorsae",
    #     "navab.ebrahimiii",
    #     "atiiye__",
    #     "animotionartz",
    #     "goldenbusinessmagazine",
    #     "shahrbanoofficial",
    #     "service_bourse",
    #     "saharyhomee",
    #     "zehnee.khalagh",
    #     "zibaegram",
    #     "zarrebein",
    #     "shervinmashayekh",
    #     "zaryhomee",
    #     "jiimomo",
    #     "mahsa_iman7700",
    #     "irdailystyle",
    #     "rojamovie",
    #     "hootan7",
    #     "tejaratnews",
    #     "sevinf4rbod",
    # ]
    success, fail = await process_usernames(usernames)

    # Retry failed usernames once more
    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("SELECT username FROM users_invalid") as cursor:
            invalid_usernames = [row[0] async for row in cursor]

    print("\n🔁 Retrying invalid usernames...")
    retry_success, retry_fail = await process_usernames(invalid_usernames)

    await export_excel()
    duration = int(time.time() - start)

    print(f"\n📊 Total Success: {success + retry_success}")
    print(f"❌ Total Fail: {fail + retry_fail}")
    print(f"🕒 Time Taken: {duration} seconds")
    print(f"📁 Excel Saved as: {EXCEL_FILE}")


if __name__ == "__main__":
    asyncio.run(main())
